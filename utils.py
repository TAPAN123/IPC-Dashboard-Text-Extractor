# utils.py

import cv2
import easyocr
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import difflib
import string

#reader = easyocr.Reader(['en'])
reader = None

ALERT_KEYWORDS = [
    "Anti-Lock Braking System", "Traction Control Indicator","Traction Control Malfunction", "Battery", "Airbag", "Brake", "Abs", "Tire Pressure low", "Door Open", "Low fuel", "Check engine", "Lane departure warning",
    "Oil pressure low", "Master Warning", "tpms"
]

def get_ocr_reader():
    global reader
    if reader is None:
        reader = easyocr.Reader(['en'], gpu=False)
    return reader

def normalize(text):
    return text.lower().translate(str.maketrans("", "", string.punctuation)).strip()

def preview_and_crop(img_path):
    image = cv2.imread(img_path)
    factor = 0.5  # or calculate based on original size
    # Resize image to 50% of assumed 1920x1080 resolution
    screen_width = int(image.shape[1] * factor)
    screen_height = int(image.shape[0] * factor)
    resized = cv2.resize(image, (screen_width, screen_height))

    roi = cv2.selectROI("Crop IPC Image (50% screen)", resized, showCrosshair=True)
    cv2.waitKey(1)
    cv2.destroyAllWindows()

    if sum(roi) == 0:
        return image  # No selection; return full image

    x, y, w, h = roi

    # Map ROI from resized to original coordinates
    x_orig = int(x * image.shape[1] / screen_width)
    y_orig = int(y * image.shape[0] / screen_height)
    w_orig = int(w * image.shape[1] / screen_width)
    h_orig = int(h * image.shape[0] / screen_height)

    cropped_img = image[y_orig:y_orig+h_orig, x_orig:x_orig+w_orig]

    # Preview cropped image
    cv2.imshow("Cropped Preview", cropped_img)
    cv2.waitKey(1000)
    cv2.destroyAllWindows()

    return cropped_img


def extract_text_from_cropped(image):
    ocr_reader = get_ocr_reader()
    result = ocr_reader.readtext(image)
    extracted_text = [normalize(text[1]) for text in result]

    matched_texts: list[str] = []
    for text in extracted_text:
        best_match = difflib.get_close_matches(text, ALERT_KEYWORDS, n=1, cutoff=0.5)
        if best_match:
            formatted = best_match[0].title()
            if formatted not in matched_texts:  # Prevent duplicates
                matched_texts.append(formatted)

    return " | ".join(matched_texts)


def save_to_excel(image_data, filename):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "IPC Alerts"
        ws.append(["Image Name", "Extracted Text", "Timestamp"])

    for image_name, text in image_data:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([image_name, text, timestamp])

    wb.save(filename)
